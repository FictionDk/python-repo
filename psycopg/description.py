import psycopg2
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

db_params = {
    'dbname': 'xxxx',
    'user': 'xxx',
    'password': 'xxx',
    'host': '192.168.110.13',
    'port': '5432'
}

def get_table_structure(cursor):
    """获取数据库中所有表的结构信息"""
    cursor.execute("""
        SELECT 
            table_schema,
            table_name,
            column_name,
            data_type,
            col_description((table_schema || '.' || table_name)::regclass, ordinal_position) AS column_comment
        FROM information_schema.columns
        WHERE table_schema NOT IN ('pg_catalog', 'information_schema')
        ORDER BY table_schema, table_name, ordinal_position;
    """)
    
    columns = cursor.fetchall()
    return columns

def get_table_comments(cursor):
    """获取数据库中所有表的注释信息"""
    cursor.execute("""
        SELECT 
            n.nspname AS table_schema,
            c.relname AS table_name,
            d.description AS table_comment
        FROM pg_class c
        LEFT JOIN pg_namespace n ON n.oid = c.relnamespace
        LEFT JOIN pg_description d ON d.objoid = c.oid AND d.objsubid = 0
        WHERE c.relkind = 'r' AND n.nspname NOT IN ('pg_catalog', 'information_schema');
    """)
    
    comments = {f"{row[1]}": f"{row[1]}-{row[2]}" for row in cursor.fetchall()}
    return comments

def main():
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        raw_data = get_table_structure(cursor)
        df = pd.DataFrame(raw_data, columns=[
            'Table Schema', 'Table Name', 'Column Name', 'Data Type', 'Column Comment'
        ])

        table_comments = get_table_comments(cursor)
        df['Table Name'] = df['Table Name'].replace(table_comments)

        df['Data Type'] = df['Data Type'].replace({
            'integer': '整型',
            'character varying': '字符',
            'boolean': '布尔',
            'timestamp without time zone': '日期时间',
            'text': "文本",
            'numeric': '长整型',
            'jsonb': 'Json文本',
            'real': '整型',
            'smallint': '整型'
        })

        # 创建一个新的 Excel 工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "Database Structure"

        # 设置特定列的宽度
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 50

        # 写入表头
        headers = ['Table Schema', 'Table Name', 'Column Name', 'Data Type', 'Column Comment']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header).alignment = Alignment(horizontal="center", vertical="center")

        # 写入数据并合并相同的 Table Name 单元格
        current_table_name = None
        start_row = 2

        for index, row in df.iterrows():
            if current_table_name != row['Table Name']:
                if current_table_name is not None:
                    ws.merge_cells(start_row=start_row, end_row=index + 1, start_column=2, end_column=2)
                current_table_name = row['Table Name']
                start_row = index + 2

            for col_num, value in enumerate(row, 1):
                ws.cell(row=index + 2, column=col_num, value=value).alignment = Alignment(wrap_text=True)

        # 处理最后一组相同 Table Name 的合并
        if current_table_name is not None:
            ws.merge_cells(start_row=start_row, end_row=len(df) + 2, start_column=2, end_column=2)

        excel_file = 'db_structure.xlsx'
        wb.save(excel_file)
        #excel_file = 'db_structure.xlsx'
        #df.to_excel(excel_file, index=False)
        print('导出成功')
    except Exception as e:
        print('导出错误:',str(e))
    finally:
        if conn:
            cursor.close()
            conn.close()

if __name__ == "__main__":
    main()


