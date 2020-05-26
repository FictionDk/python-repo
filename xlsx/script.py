# -*- coding: utf-8 -*-

import openpyxl
import os
import shutil

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

class ArgsNotNull(RuntimeError):
    pass

def get_full_filename(directory,filename=None):
    '''获取当前执行目录下的指定路径下的文件名
    Args:
        filename: 文件名,可为空,返回文件夹全路径
        directory: 目录名称
    Returns:
        返回文件的全路径
    Raises:
        ArgsNotNull: 参数不能为空
    '''
    if directory is None:
        raise ArgsNotNull("directory can not be None")
    dir_name = os.path.join(BASE_DIR,directory)

    # 确保目录存在
    if not os.path.exists(dir_name):
        os.makedirs(dir_name)

    if filename is None:
        return dir_name
    full_name = os.path.join(dir_name,filename)
    # 确保文件存在
    if not os.path.isfile(full_name):
        with open(full_name,'a+',encoding="utf-8") as f:
            f.close()
    return full_name

def get_filename_in_dir(directory):
    '''获取当前执行目录下的指定路径下的文件名列表
    Args:
        directory: 目录名称

    Returns:
        返回文件的全路径
    '''
    dir_full_path = get_full_filename(directory)
    return os.listdir(dir_full_path)

def build_md_table():
    cells = get_bag_conf_data()
    for cell in cells:
        print("%s | %s | %s | %s" %(cell['地区'],cell['规格'],cell['名称'],cell['尺寸']))


def get_bag_conf_data():
    file_name_list = get_filename_in_dir("docs")
    cells = []
    for file_name in file_name_list:
        area_name = file_name.replace('.xlsx','')
        file_name = get_full_filename("docs",file_name)
        print(">> %s starting ..."%area_name)
        workbook = openpyxl.load_workbook(file_name,read_only=True)
        ws = workbook.active
        rows = ws.rows
        for i,row in enumerate(rows):
            cells += _formate_date(row,area_name)
            print("i = %d len = %d" % (i,len(cells)))
            print(cells)
        print(">> %s end ..."%area_name)
    return cells

def _formate_date(row,area_name):
    row_contain_data = lambda row: len(row) == 11 and str(row[0].value).isdigit()
    cells = []
    # print(row_contain_data(row),",len=",len(row),",row[0]=",str(row[0]))
    if row_contain_data(row):
        bag = {}
        bag['50ml'] = row[2].value
        bag['100ml'] = row[3].value
        bag['150ml'] = row[4].value
        bag['200ml'] = row[5].value
        bag['250ml'] = row[6].value
        bag['0.5u'] = row[7].value
        bag['1.0u'] = row[8].value
        bag['1.5u'] = row[9].value
        bag['2.0u'] = row[10].value
        for j,col in enumerate(row):
            if j > 1 and col.value is not None:
                cell = {}
                if j == 2:
                    cell['规格'] = '50ml'
                elif j == 3:
                    cell['规格'] = '100ml'
                elif j == 4:
                    cell['规格'] = '150ml'
                elif j == 5:
                    cell['规格'] = '200ml'
                elif j == 6:
                    cell['规格'] = '250ml'
                elif j == 7:
                    cell['规格'] = '0.5u'
                elif j == 8:
                    cell['规格'] = '1.0u'
                elif j == 9:
                    cell['规格'] = '1.5u'
                elif j == 10:
                    cell['规格'] = '2.0u'
                cell['地区'] = area_name
                cell['名称'] = row[1].value
                cell['尺寸'] = col.value.replace('*',' * ')
                cells.append(cell)
    return cells

def _print_row(i,row):
    for j,col in enumerate(row):
        print("row[%d] col[%d] = %s" %(i,j,str(col.value)))

def _get_pic_uri(own_machine=True):
    """
    D://Doc//OneDrive//Pictures//
    E://OneDrive//Pictures//
    """
    if own_machine:
        return os.path.join('D:',os.path.sep,'Doc','OneDrive')
    else:
        return os.path.join('E:',os.path.sep,'OneDrive')

def _get_pic_files():
    path = _get_pic_uri()
    files = os.listdir(path)
    print(path)
    pic_files = []
    for i in range(len(files)):
        if _is_pic_file(files[i],path):
            pic_files.append(files[i])
    return pic_files

def _is_pic_file(filename,path):
    full_filename = os.path.join(path,filename)
    if os.path.isfile(full_filename) and 'Pictures' == filename[0:8]:
        return True
    else:
        return False

def _get_full_filename(filename,dir):
    dir_name = os.path.join(os.getcwd(),dir)
    if not os.path.exists(dir_name):
        os.makedirs(dir_name)
    full_name = os.path.join(dir_name,filename)
    # 确保被创建
    if not os.path.isfile(full_name):
        with open(full_name,'a+',encoding="utf-8") as f:
            f.close()
    return full_name

def _get_content():
    content = []

    old_path = _get_pic_uri()
    new_path = os.path.join(old_path,"Pictures")
    if not os.path.exists(new_path):
        os.makedirs(new_path)

    pic_files = _get_pic_files()
    for pic in pic_files:
        pic_content = [os.path.join(old_path,pic),os.path.join(new_path,pic.replace('Pictures',''))]
        content.append(pic_content)

    return content

def build_xlsx():
    xlsx_name = _get_full_filename("test.xlsx","docs")
    workbook = openpyxl.Workbook()
    ws = workbook.active

    pics = _get_content()
    for pic in pics:
        ws.append(pic)
    workbook.save(xlsx_name)

def files_move():
    xlsx_name = _get_full_filename("test.xlsx","docs")
    workbook = openpyxl.load_workbook(xlsx_name,read_only=True)
    ws = workbook.active

    rows = ws.rows
    for row in rows:
        print(row[0].value,"->",row[1].value)
        try:
            shutil.move(row[0].value,row[1].value)
        except FileNotFoundError as e:
            print("FileNotFoundError:",e)

def main():
    # build_xlsx()
    # files_move()
    build_md_table()

if __name__ == '__main__':
    main()
